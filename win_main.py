from Forms import winMain, dlgAbout
from win_settings import SettingsDlg
from win_about import AboutDlg
from general_functions import *
from ExGrid import ExGrid
from views import VIEWS
import dbconnection
import openpyxl
import wx.adv


class MainWindow(winMain):
    def __init__(self, parent):
        super().__init__(parent)
        set_layout(self)
        self.snap = None

        # add all grids
        D = dbconnection.Database(connection_string())
        for view, readonly, title in VIEWS:
            self.append_grid(title, D.GetFullTable(view), readonly)
        D.Close()

        # append calculated columns
        grid = self.nBook.GetPage(8)  # 2W Tap changers
        grid.AppendCols(1)
        grid.SetColLabelValue(grid.NumberCols - 1, 'TTR')

        grid = self.nBook.GetPage(9)  # 3W Tap changers
        grid.AppendCols(1)
        grid.SetColLabelValue(grid.NumberCols - 1, 'TTR')

        del grid

    # create a grid of table data and add it to notebook as a page
    def append_grid(self, title, table=None, readonly_cols=()):
        grid = ExGrid(self.nBook, readonly_cols)
        grid.CreateGrid(0, table.ColsCount)

        for c in range(table.ColsCount):
            grid.SetColLabelValue(c, table.Columns[c])

        # set column types
        c = 0
        for t in table.Types:
            if t is bool:
                grid.SetColFormatBool(c)
            elif t is float:
                grid.SetColFormatFloat(c)
            elif t is int:
                grid.SetColFormatNumber(c)
            c += 1

        # fill the grid with table values
        grid.FillValues(table.Data)

        # remove column 'ID' if exists
        for c in range(grid.NumberCols):
            if grid.ColumnLabels[c] == 'ID':
                grid.DeleteCols(c)
                break

        grid.AutoSizeColumns()
        self.nBook.AddPage(grid, text=title)
        return grid

    # show Settings dialog modally then destroys it. every action takes place
    # in the dialg init() and close() methods.
    def onSettings(self, event):
        with SettingsDlg(self) as dlg:
            dlg.ShowModal()

    # exit the application
    def onExit(self, event):
        self.Close()

    def onLoadSnap(self, event):
        with wx.FileDialog(self, 'Open snapshot file', defaultDir=snap_dir(), style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            else:
                snap_file_name = dlg.GetPath()

        with Wait('Finished loading snapshot'):
            self.snap = read_snap(snap_file_name)
            for i in range(self.nBook.GetPageCount()):
                grid = self.nBook.GetPage(i)
                inv = grid.ApplySnap(self.snap)

    # temp code for Mehdi Kabiri
    def on_apply_snapp(self, event):
        if self.snap is None:
            wx.MessageBox('No snapshot loaded. Please load a snapshot first.', 'Error', wx.OK)
            return

        with Wait('Finished putting snapshot in database'):
            snp = self.snap
            snp_to_db = []
            validity = {0: 'i', 1: 'v', 2: 'M', 3: 'm'}
            for obj in snp:
                val, typ = snp[obj]
                val_str = validity[typ] + str(val)
                snp_to_db.append((obj, val_str))
            snp_to_db = tuple(snp_to_db)

            try:
                D = dbconnection.Database(connection_string())
                D.RunQuery("DELETE FROM tbl_snap")
                grp = len(snp_to_db) // 1000
                for g in range(grp + 1):
                    values_str = str(snp_to_db[g * 1000:(g + 1) * 1000])
                    D.RunQuery("INSERT INTO tbl_snap VALUES %s" % values_str[1:-1])
            finally:
                D.Close()

    def onExport(self, event):
        with wx.FileDialog(self, 'Save as Excel file',wildcard="Excel files (*.xlsx)|*.xlsx", style=wx.FD_SAVE) as dlg:
            if dlg.ShowModal() == wx.ID_CANCEL:
                return
            else:
                workbook_filename = dlg.GetPath()
                wb = openpyxl.Workbook()

        with Wait('Finished exporting to excel document'):
            for p in range(self.nBook.GetPageCount()):
                sh_name = self.nBook.GetPageText(p)
                grid = self.nBook.GetPage(p)
                ws = wb.create_sheet(sh_name)
                for c, label in enumerate(grid.ColumnLabels, 1):
                    ws.cell(1, c).value = label

                for r, row in enumerate(grid.CellValues, 2):
                    for c, value in enumerate(row, 1):
                        ws.cell(r, c).value = convert_string(value)
            wb.remove_sheet(wb.get_sheet_by_name(wb.sheetnames[0]))
            wb.save(workbook_filename)

    def onAbout(self, event):
        dlg = AboutDlg(self)
        dlg.ShowModal()
        dlg.Destroy()
